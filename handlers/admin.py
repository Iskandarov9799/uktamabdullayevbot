import io
import openpyxl
from aiogram import Router, F, Bot
from aiogram.types import Message, CallbackQuery, BufferedInputFile, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext

from database.db import (get_all_users, get_full_stats, add_question, count_questions, delete_all_questions)
from keyboards.keyboards import (admin_keyboard, cancel_keyboard, main_menu_keyboard, subject_keyboard,
                                 addq_category_keyboard, addq_topic_keyboard, addq_grade_keyboard,
                                 addq_difficulty_keyboard, correct_answer_keyboard, skip_image_keyboard)
from states import AdminStates
from config import config

router = Router()

def is_admin(message: Message) -> bool:
    return message.from_user.id in config.ADMIN_IDS


@router.message(F.text == "👥 Foydalanuvchilar")
async def admin_users(message: Message):
    if not is_admin(message):
        return
    users = await get_all_users()
    text  = f"👥 <b>Foydalanuvchilar: {len(users)} ta</b>\n\n"
    for u in users[:20]:
        icon     = "✅" if u.is_registered else "👤"
        name     = u.full_name or "Noma'lum"
        phone    = u.phone_number or "—"
        username = "@" + u.username if u.username else "—"
        text += f"{icon} {name} | {phone} | {username}\n"
    if len(users) > 20:
        text += f"\n... va yana {len(users) - 20} ta"
    await message.answer(text, parse_mode="HTML")


@router.message(F.text == "📥 Excel eksport")
async def admin_export(message: Message):
    if not is_admin(message):
        return
    users = await get_all_users()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"
    ws.append(["#", "Telegram ID", "Ism", "Username", "Telefon", "Ro'yxat", "Sana"])
    for i, u in enumerate(users, 1):
        ws.append([
            i, u.telegram_id, u.full_name or "", u.username or "",
            u.phone_number or "", "Ha" if u.is_registered else "Yo'q",
            str(u.registered_at)[:16] if u.registered_at else "",
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 2, 40
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await message.answer_document(
        document=BufferedInputFile(buf.read(), filename="users.xlsx"),
        caption=f"📥 Jami: <b>{len(users)}</b> ta",
        parse_mode="HTML"
    )


@router.message(F.text == "📢 Broadcast")
async def admin_broadcast_start(message: Message, state: FSMContext):
    if not is_admin(message):
        return
    await message.answer(
        "📢 <b>Broadcast</b>\n\nXabarni yozing:",
        reply_markup=cancel_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.broadcast_message)


@router.message(AdminStates.broadcast_message, F.text == "❌ Bekor qilish")
async def broadcast_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("❌ Bekor qilindi.", reply_markup=admin_keyboard())


@router.message(AdminStates.broadcast_message)
async def admin_broadcast_send(message: Message, state: FSMContext, bot: Bot):
    await state.clear()
    users   = await get_all_users()
    success = 0
    failed  = 0
    await message.answer(f"⏳ {len(users)} ta foydalanuvchiga yuborilmoqda...")
    for user in users:
        try:
            if message.photo:
                await bot.send_photo(user.telegram_id, photo=message.photo[-1].file_id, caption=message.caption or "")
            elif message.video:
                await bot.send_video(user.telegram_id, video=message.video.file_id, caption=message.caption or "")
            else:
                await bot.send_message(user.telegram_id, text=message.text or "")
            success += 1
        except Exception:
            failed += 1
    await message.answer(
        f"📢 <b>Broadcast tugadi!</b>\n\n✅ {success}\n❌ {failed}",
        reply_markup=admin_keyboard(),
        parse_mode="HTML"
    )


@router.message(F.text == "➕ Savol qo'shish")
async def add_question_start(message: Message, state: FSMContext):
    if not is_admin(message):
        return
    await message.answer(
        "➕ <b>Yangi savol qo'shish</b>\n\nFanni tanlang:",
        reply_markup=subject_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.add_subject)


@router.callback_query(F.data.startswith("addq:subject:"))
async def addq_subject(callback: CallbackQuery, state: FSMContext):
    subject = callback.data.split(":")[2]
    await state.update_data(subject=subject)
    labels = {'jahon': '🌍 Jahon tarixi', 'uzbekiston': "🇺🇿 O'zbekiston tarixi"}
    label  = labels.get(subject, subject)
    await callback.message.edit_text(
        f"📁 <b>{label} — Kategoriya tanlang:</b>",
        reply_markup=addq_category_keyboard(subject),
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.add_category)
    await callback.answer()


@router.callback_query(F.data.startswith("addq:cat:"))
async def addq_category(callback: CallbackQuery, state: FSMContext):
    category = callback.data.split(":")[2]
    await state.update_data(category=category)
    data    = await state.get_data()
    subject = data.get('subject')

    if category == 'mavzu':
        await callback.message.edit_text(
            "📌 <b>Mavzuni tanlang:</b>",
            reply_markup=addq_topic_keyboard(subject),
            parse_mode="HTML"
        )
        await state.set_state(AdminStates.add_subcategory)

    elif category == 'sinf':
        await callback.message.edit_text(
            "🏫 <b>Sinfni tanlang:</b>",
            reply_markup=addq_grade_keyboard(subject),
            parse_mode="HTML"
        )
        await state.set_state(AdminStates.add_subcategory)

    elif category == 'attestation':
        await state.update_data(subcategory=None, difficulty=None, is_attestation=True)
        cnt = await count_questions(subject=subject, is_attestation=True)
        await callback.message.edit_text(
            f"🎓 <b>Atestatsiya savoli</b>\n\nHozir {cnt} ta bor.\n"
            f"Tartib raqamini yozing (masalan: <code>{cnt+1}</code>):",
            parse_mode="HTML"
        )
        await state.set_state(AdminStates.add_order_num)

    else:
        await callback.message.edit_text(
            "🎯 <b>Qiyinlik darajasini tanlang:</b>",
            reply_markup=addq_difficulty_keyboard(),
            parse_mode="HTML"
        )
        await state.set_state(AdminStates.add_difficulty)

    await callback.answer()


@router.callback_query(F.data.startswith("addq:topic:"))
async def addq_topic(callback: CallbackQuery, state: FSMContext):
    topic = callback.data.split(":")[2]
    await state.update_data(subcategory=topic, is_attestation=False)
    data    = await state.get_data()
    subject = data.get('subject', 'jahon')
    topics  = config.JAHON_TOPICS if subject == 'jahon' else config.UZBEKISTON_TOPICS
    label   = topics.get(topic, topic)
    await callback.message.edit_text(
        f"📌 <b>{label}</b>\n\nQiyinlik darajasini tanlang:",
        reply_markup=addq_difficulty_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.add_difficulty)
    await callback.answer()


@router.callback_query(F.data.startswith("addq:grade:"))
async def addq_grade(callback: CallbackQuery, state: FSMContext):
    grade = callback.data.split(":")[2]
    await state.update_data(subcategory=grade, is_attestation=False)
    data    = await state.get_data()
    subject = data.get('subject', 'jahon')
    grades  = config.JAHON_GRADES if subject == 'jahon' else config.UZBEKISTON_GRADES
    await callback.message.edit_text(
        f"🏫 <b>{grades.get(grade, grade)}</b>\n\nQiyinlik darajasini tanlang:",
        reply_markup=addq_difficulty_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.add_difficulty)
    await callback.answer()


@router.callback_query(F.data.startswith("addq:diff:"))
async def addq_difficulty(callback: CallbackQuery, state: FSMContext):
    difficulty = callback.data.split(":")[2]
    await state.update_data(difficulty=difficulty, is_attestation=False)
    DIFF = {'easy': '🟢 Oson', 'medium': "🟡 O'rta", 'hard': '🔴 Qiyin'}
    await callback.message.edit_text(
        f"✅ Qiyinlik: <b>{DIFF.get(difficulty)}</b>",
        parse_mode="HTML"
    )
    await callback.message.answer("📸 Rasm yuboring yoki o'tkazib yuboring:", reply_markup=skip_image_keyboard())
    await state.set_state(AdminStates.add_image)
    await callback.answer()


@router.message(AdminStates.add_order_num)
async def addq_order_num(message: Message, state: FSMContext):
    try:
        order_num = int(message.text.strip())
        await state.update_data(order_num=order_num)
    except ValueError:
        await message.answer("❌ Raqam kiriting!")
        return
    await message.answer("📸 Rasm yuboring yoki o'tkazib yuboring:", reply_markup=skip_image_keyboard())
    await state.set_state(AdminStates.add_image)


@router.message(AdminStates.add_image, F.photo)
async def addq_image(message: Message, state: FSMContext):
    await state.update_data(image_file_id=message.photo[-1].file_id)
    await message.answer("✅ Rasm saqlandi.\n\n📝 Savol matnini yozing:", reply_markup=cancel_keyboard())
    await state.set_state(AdminStates.add_text)


@router.message(AdminStates.add_image, F.text == "⏭ Rasmisiz davom etish")
async def addq_skip_image(message: Message, state: FSMContext):
    await state.update_data(image_file_id=None)
    await message.answer("📝 Savol matnini yozing:", reply_markup=cancel_keyboard())
    await state.set_state(AdminStates.add_text)


@router.message(AdminStates.add_text)
async def addq_text(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("❌ Bekor qilindi.", reply_markup=admin_keyboard())
        return
    await state.update_data(question_text=message.text)
    await message.answer("🅰 A variantini yozing:")
    await state.set_state(AdminStates.add_a)


@router.message(AdminStates.add_a)
async def addq_a(message: Message, state: FSMContext):
    await state.update_data(option_a=message.text)
    await message.answer("🅱 B variantini yozing:")
    await state.set_state(AdminStates.add_b)


@router.message(AdminStates.add_b)
async def addq_b(message: Message, state: FSMContext):
    await state.update_data(option_b=message.text)
    await message.answer("🅲 C variantini yozing:")
    await state.set_state(AdminStates.add_c)


@router.message(AdminStates.add_c)
async def addq_c(message: Message, state: FSMContext):
    await state.update_data(option_c=message.text)
    await message.answer("🅳 D variantini yozing:")
    await state.set_state(AdminStates.add_d)


@router.message(AdminStates.add_d)
async def addq_d(message: Message, state: FSMContext):
    await state.update_data(option_d=message.text)
    await message.answer("✅ <b>To'g'ri javobni tanlang:</b>", reply_markup=correct_answer_keyboard(), parse_mode="HTML")
    await state.set_state(AdminStates.add_correct)


@router.callback_query(F.data.startswith("addq:correct:"))
async def addq_correct(callback: CallbackQuery, state: FSMContext):
    correct = callback.data.split(":")[2]
    data    = await state.get_data()
    await state.clear()

    await add_question(
        subject=data['subject'],
        category=data['category'],
        question_text=data['question_text'],
        option_a=data['option_a'],
        option_b=data['option_b'],
        option_c=data['option_c'],
        option_d=data['option_d'],
        correct_answer=correct,
        subcategory=data.get('subcategory'),
        difficulty=data.get('difficulty'),
        is_attestation=data.get('is_attestation', False),
        order_num=data.get('order_num'),
        image_file_id=data.get('image_file_id'),
    )

    labels = {'jahon': '🌍 Jahon tarixi', 'uzbekiston': "🇺🇿 O'zbekiston tarixi"}
    subj_label  = labels.get(data['subject'], data['subject'])
    subcategory = data.get('subcategory') or '—'
    difficulty  = data.get('difficulty') or '—'

    await callback.message.edit_text(
        f"✅ <b>Savol qo'shildi!</b>\n\n"
        f"📜 Fan: {subj_label}\n"
        f"📁 Kategoriya: {data['category']}\n"
        f"🔑 Subcategory: {subcategory}\n"
        f"🎯 Qiyinlik: {difficulty}\n"
        f"✅ To'g'ri javob: <b>{correct}</b>",
        parse_mode="HTML"
    )
    await callback.message.answer("Admin panel:", reply_markup=admin_keyboard())
    await callback.answer()


@router.callback_query(F.data == "addq:cancel")
async def addq_cancel(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("❌ Bekor qilindi.")
    await callback.message.answer("Admin panel:", reply_markup=admin_keyboard())
    await callback.answer()


@router.message(F.text == "🗑 Savollarni o'chirish")
async def delete_questions_confirm(message: Message):
    if not is_admin(message):
        return
    cnt = await count_questions()
    await message.answer(
        f"⚠️ <b>Diqqat!</b>\n\nHozir bazada <b>{cnt} ta</b> savol bor.\n"
        f"Barchasini o'chirishni tasdiqlaysizmi?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="🗑 Ha, o'chirish", callback_data="admin:delete_all_q"),
            InlineKeyboardButton(text="❌ Yo'q",          callback_data="admin:cancel_delete"),
        ]]),
        parse_mode="HTML"
    )


@router.callback_query(F.data == "admin:delete_all_q")
async def delete_questions_execute(callback: CallbackQuery):
    deleted = await delete_all_questions()
    await callback.message.edit_text(f"🗑 <b>{deleted} ta savol o'chirildi!</b>", parse_mode="HTML")
    await callback.message.answer("Admin panel:", reply_markup=admin_keyboard())
    await callback.answer()


@router.callback_query(F.data == "admin:cancel_delete")
async def delete_questions_cancel(callback: CallbackQuery):
    await callback.message.edit_text("❌ Bekor qilindi.")
    await callback.answer()


@router.message(F.text == "📤 Excel import")
async def excel_import_start(message: Message):
    if not is_admin(message):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savollar"
    headers = ["subject", "category", "subcategory", "difficulty",
               "is_attestation", "order_num", "question", "a", "b", "c", "d", "correct"]
    ws.append(headers)
    from openpyxl.styles import Font, PatternFill
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    for i, w in enumerate([12, 12, 14, 12, 14, 10, 50, 25, 25, 25, 25, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    examples = [
        ["jahon",      "mavzu",       "qadimgi",  "easy",   "FALSE", "",  "Qadimgi Misr podshohlari qanday atalgan?",         "Fir'avn",             "Imperator",           "Xon",                   "Shoh",                   "A"],
        ["jahon",      "aralash",     "",          "medium", "FALSE", "",  "Birinchi jahon urushi qachon boshlangan?",          "1912",                "1914",                "1916",                  "1918",                   "B"],
        ["uzbekiston", "sinf",        "9",         "hard",   "FALSE", "",  "O'zbekiston mustaqilligini qachon qo'lga kiritdi?", "1990",                "1991",                "1992",                  "1993",                   "B"],
        ["uzbekiston", "attestation", "",          "",       "TRUE",  "1", "O'zbekiston Respublikasi poytaxti qaysi shahar?",   "Samarqand",           "Toshkent",            "Buxoro",                "Namangan",               "B"],
    ]
    for row in examples:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await message.answer_document(
        document=BufferedInputFile(buf.read(), filename="tarix_savollar_shablon.xlsx"),
        caption=(
            "📋 <b>Excel shablon</b>\n\n"
            "1️⃣ Yuklab oling\n"
            "2️⃣ To'ldiring\n"
            "3️⃣ Botga yuboring\n\n"
            "📌 subject: <code>jahon</code> yoki <code>uzbekiston</code>"
        ),
        parse_mode="HTML"
    )


@router.message(F.document)
async def excel_import_upload(message: Message):
    if not is_admin(message):
        return
    doc = message.document
    if not doc.file_name or not doc.file_name.endswith('.xlsx'):
        return
    await message.answer("⏳ Fayl o'qilmoqda...")
    bot: Bot = message.bot
    file = await bot.get_file(doc.file_id)
    buf  = io.BytesIO()
    await bot.download_file(file.file_path, buf)
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    VALID_SUBJECTS = {'jahon', 'uzbekiston'}
    VALID_CATS     = {'mavzu', 'aralash', 'sinf', 'attestation'}
    VALID_CORR     = {'A', 'B', 'C', 'D'}
    added   = 0
    skipped = 0
    errors  = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        try:
            subject        = str(row[0]  or '').strip().lower()
            category       = str(row[1]  or '').strip().lower()
            subcategory    = str(row[2]  or '').strip() or None
            difficulty     = str(row[3]  or '').strip().lower()
            is_attestation = str(row[4]  or '').strip().upper() == 'TRUE'
            order_num      = int(row[5]) if row[5] else None
            question_text  = str(row[6]  or '').strip()
            option_a       = str(row[7]  or '').strip()
            option_b       = str(row[8]  or '').strip()
            option_c       = str(row[9]  or '').strip()
            option_d       = str(row[10] or '').strip()
            correct        = str(row[11] or '').strip().upper()

            if subject not in VALID_SUBJECTS:
                errors.append(f"Qator {row_num}: subject '{subject}' noto'g'ri (jahon yoki uzbekiston)")
                skipped += 1; continue
            if category not in VALID_CATS:
                errors.append(f"Qator {row_num}: category noto'g'ri")
                skipped += 1; continue
            if not question_text:
                errors.append(f"Qator {row_num}: savol bo'sh")
                skipped += 1; continue
            if correct not in VALID_CORR:
                errors.append(f"Qator {row_num}: correct noto'g'ri")
                skipped += 1; continue
            if not all([option_a, option_b, option_c, option_d]):
                errors.append(f"Qator {row_num}: variantlar to'liq emas")
                skipped += 1; continue

            await add_question(
                subject=subject, category=category, subcategory=subcategory,
                difficulty=difficulty or None, is_attestation=is_attestation,
                order_num=order_num, question_text=question_text,
                option_a=option_a, option_b=option_b,
                option_c=option_c, option_d=option_d,
                correct_answer=correct
            )
            added += 1
        except Exception as e:
            errors.append(f"Qator {row_num}: {e}")
            skipped += 1

    text = (
        f"✅ <b>Import tugadi!</b>\n\n"
        f"✅ Qo'shildi: <b>{added} ta</b>\n"
        f"❌ O'tkazildi: <b>{skipped} ta</b>"
    )
    if errors:
        err_text = "\n".join(errors[:10])
        if len(errors) > 10:
            err_text += f"\n... va yana {len(errors)-10} ta"
        text += f"\n\n⚠️ <b>Xatolar:</b>\n<code>{err_text}</code>"
    await message.answer(text, parse_mode="HTML", reply_markup=admin_keyboard())